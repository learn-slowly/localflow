#!/usr/bin/env python3
"""NEC xlsx에서 진주시 읍면동별 비례대표 투표 결과 파싱 → elections JSON에 추가"""

import json
import os
import openpyxl

BASE = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE, "..", "public", "data", "elections")

SOURCES = [
    {
        "sgId": "20240410",
        "label": "2024 총선",
        "path": "src/data/nec/2024총선개표결과/2. 개표단위별 개표결과(비례대표) -전국.xlsx",
        "city": "진주시",
    },
    {
        "sgId": "20200415",
        "label": "2020 총선",
        "path": "src/data/nec/2020제21대 국회의원선거(재보궐 포함) 투표구별 개표결과/비례대표/16경남/개표상황(투표구별)_진주시.xlsx",
        "city": "진주시",
    },
]


def parse_int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    try:
        return int(float(s))
    except:
        return 0


def detect_format(all_rows):
    """전국 파일(열 4개: 시도/구시군/읍면동/투표구) vs 시군구 파일(열 2개: 읍면동/투표구) 구분"""
    for row in all_rows[:10]:
        vals = list(row)
        if len(vals) > 1 and str(vals[0] or "").strip() == "읍면동명":
            return "city"  # 시군구별 파일
        if len(vals) > 3 and str(vals[0] or "").strip() == "시도명":
            return "national"  # 전국 파일
    return "national"


def parse_xlsx(path, city_name):
    """xlsx에서 읍면동별 소계 행의 정당별 득표 파싱"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    all_rows = list(ws.iter_rows(values_only=True))
    fmt = detect_format(all_rows)

    if fmt == "city":
        # 시군구 파일: [읍면동, 투표구, 선거인수, 투표수, 정당1, 정당2, ...]
        emd_col, tpg_col, voters_col, turnout_col, party_start = 0, 1, 2, 3, 4
    else:
        # 전국 파일: [시도, 구시군, 읍면동, 투표구, 선거인수, 투표수, 정당1, ...]
        emd_col, tpg_col, voters_col, turnout_col, party_start = 2, 3, 4, 5, 6

    # 정당 이름 행 찾기
    parties = []
    for row in all_rows[:10]:
        vals = list(row)
        if len(vals) > party_start and vals[party_start]:
            v = str(vals[party_start]).strip()
            if any(kw in v for kw in ["더불어", "국민", "미래", "한국", "민주", "민생", "정의"]):
                parties = [str(x).strip() for x in vals[party_start:] if x and str(x).strip()]
                break

    if not parties:
        print(f"  정당 헤더 못 찾음: {path}")
        return None, []

    print(f"  정당 {len(parties)}개: {parties[:5]}...")

    SKIP_NAMES = {"합계", "거소·선상투표", "관외사전투표", "재외투표",
                  "재외투표(공관)", "재외투표(재외)", "잘못 투입·구분된 투표지"}

    dong_results = []
    in_city = False

    for row in all_rows:
        vals = list(row)
        if len(vals) <= party_start:
            continue

        if fmt == "national":
            sgg = str(vals[1] or "").strip()
            if city_name in sgg:
                in_city = True
            elif in_city and sgg and city_name not in sgg:
                break
            if not in_city:
                continue

        emd = str(vals[emd_col] or "").strip()
        tpg = str(vals[tpg_col] or "").strip()

        if emd in SKIP_NAMES or not emd:
            continue

        if tpg != "소계":
            continue

        voters = parse_int(vals[voters_col])
        turnout = parse_int(vals[turnout_col])
        party_votes = {}
        valid = 0
        for j, party in enumerate(parties):
            idx = party_start + j
            v = parse_int(vals[idx]) if idx < len(vals) else 0
            if v > 0:
                party_votes[party] = v
                valid += v
        dong_results.append({
            "dong": emd,
            "voters": voters,
            "turnout": turnout,
            "valid": valid,
            "parties": party_votes,
        })

    # 시 전체 합산
    total = {"voters": 0, "turnout": 0, "valid": 0, "invalid": 0, "parties": {}}
    for d in dong_results:
        total["voters"] += d["voters"]
        total["turnout"] += d["turnout"]
        total["valid"] += d["valid"]
        for party, votes in d["parties"].items():
            total["parties"][party] = total["parties"].get(party, 0) + votes

    total_parties = [{"party": p, "votes": v} for p, v in sorted(total["parties"].items(), key=lambda x: -x[1])]
    total["parties"] = total_parties

    return total, dong_results


def main():
    json_path = os.path.join(DATA_DIR, "48170-elections.json")
    with open(json_path, "r", encoding="utf-8") as f:
        elections = json.load(f)

    for src in SOURCES:
        path = os.path.join(BASE, "..", src["path"])
        print(f"\n=== {src['label']} 비례대표 ===")

        if not os.path.exists(path):
            print(f"  파일 없음: {path}")
            continue

        total, dong_results = parse_xlsx(path, src["city"])

        if not dong_results:
            print("  데이터 없음")
            continue

        print(f"  {len(dong_results)}개 읍면동")
        top = sorted(total["parties"], key=lambda p: -p["votes"])[:5]
        for p in top:
            pct = (p["votes"] / total["valid"] * 100) if total["valid"] else 0
            print(f"    {p['party']}: {p['votes']:,}표 ({pct:.1f}%)")

        # 기존 데이터에 추가
        for elec in elections:
            if elec.get("sgId") == src["sgId"] and "총선" in elec.get("label", ""):
                elec["proportional"] = total
                elec["proportionalByDong"] = dong_results
                print(f"  → '{elec['label']}'에 추가 완료")
                break

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(elections, f, ensure_ascii=False, indent=2)
    print(f"\n저장: {json_path}")


if __name__ == "__main__":
    main()
