#!/usr/bin/env python3
"""NEC(선관위) xlsx 파일에서 경남 읍면동별 선거결과를 추출하여
각 도시의 elections/local-elections JSON에 dongResults를 추가"""

import json
import os
import openpyxl

NEC_DIR = os.path.join(os.path.dirname(__file__), "..", "src", "data", "nec")
EXTRACT_DIR = "/tmp/nec-extract"
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "public", "data", "elections")

CITY_CODE_MAP = {
    "창원시의창구": "48121", "창원시성산구": "48123", "창원시마산합포구": "48125",
    "창원시마산회원구": "48127", "창원시진해구": "48129",
    "진주시": "48170", "통영시": "48220", "사천시": "48240", "김해시": "48250",
    "밀양시": "48270", "거제시": "48310", "양산시": "48330",
    "의령군": "48720", "함안군": "48730", "창녕군": "48740", "고성군": "48820",
    "남해군": "48840", "하동군": "48850", "산청군": "48860", "함양군": "48870",
    "거창군": "48880", "합천군": "48890",
}

SKIP_DONGS = {"합계", "거소·선상투표", "거소투표", "관외사전투표", "재외투표", "", "계"}


def clean_num(v):
    """숫자 문자열 정리"""
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    try:
        return int(float(s))
    except:
        return 0


def parse_candidates_from_header(row):
    """헤더 행에서 후보자 정보 추출 (정당\n이름 형태)"""
    candidates = []
    for cell in row:
        if cell and "\n" in str(cell):
            # _x000D_ 제거 (2018 형식)
            text = str(cell).replace("_x000D_", "").replace("\r", "")
            parts = text.split("\n")
            party = parts[0].strip()
            name = parts[1].strip() if len(parts) > 1 else ""
            if name:
                candidates.append({"name": name, "party": party})
    return candidates


def make_dong_result(dong_name, district, voters, turnout, candidates, votes_list):
    """dongResult 항목 생성"""
    valid = sum(votes_list)
    rates = {}
    votes_map = {}
    for cand, v in zip(candidates, votes_list):
        if valid > 0:
            rates[cand["name"]] = round(v / valid * 100, 1)
        votes_map[cand["name"]] = v
    return {
        "dong": dong_name,
        "district": district,
        "voters": voters,
        "turnout": turnout,
        "rates": rates,
        "votes": votes_map,
    }


# ===== 대선 파싱 =====

def parse_presidential(filepath, label, date):
    """대선 xlsx 파싱 — 읍면동 소계 추출"""
    print(f"  파싱: {label}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # 헤더에서 후보자 추출
    candidates = []
    header_row = None
    cand_start_col = None

    rows = list(ws.iter_rows(max_col=30, values_only=True))
    wb.close()

    # 후보자 행 찾기
    for i, row in enumerate(rows):
        r = [str(v or "") for v in row]
        # 후보자 정보가 있는 행 (정당\n이름)
        has_cand = sum(1 for c in row if c and "\n" in str(c))
        if has_cand >= 3:
            candidates = parse_candidates_from_header(row)
            header_row = i
            # 후보자 시작 열 찾기
            for j, c in enumerate(row):
                if c and "\n" in str(c):
                    cand_start_col = j
                    break
            break
        # 헤더 행 (시도명, 구시군명...)
        if "시도" in r[0] or "시도명" in r[0]:
            # 다음 행이 후보자
            if i + 1 < len(rows):
                next_row = rows[i + 1]
                has_cand2 = sum(1 for c in next_row if c and "\n" in str(c))
                if has_cand2 >= 3:
                    candidates = parse_candidates_from_header(next_row)
                    header_row = i + 1
                    for j, c in enumerate(next_row):
                        if c and "\n" in str(c):
                            cand_start_col = j
                            break
                    break

    if not candidates:
        # 첫 행에 정당\n이름이 있는 경우 (2022 대선)
        for i, row in enumerate(rows[:3]):
            cands = parse_candidates_from_header(row)
            if len(cands) >= 3:
                candidates = cands
                header_row = i
                for j, c in enumerate(row):
                    if c and "\n" in str(c):
                        cand_start_col = j
                        break
                break

    if not candidates:
        print(f"    후보자 정보를 찾을 수 없음")
        return {}

    print(f"    후보자 {len(candidates)}명: {', '.join(c['name'] for c in candidates)}")

    # 경남 시군구별 읍면동 추출
    city_dong_results = {}
    current_city = None
    data_start = header_row + 1 if header_row is not None else 2
    num_cands = len(candidates)

    in_gyeongnam = False

    for row in rows[data_start:]:
        r = [str(v or "").strip() for v in row]

        sido = r[0] if r[0] else None
        sigun = r[1] if len(r) > 1 and r[1] else None
        dong = r[2] if len(r) > 2 and r[2] else None
        tpg = r[3] if len(r) > 3 and r[3] else None

        # 시도 전환 감지
        if sido and "경상남도" in sido:
            in_gyeongnam = True
        elif sido and sido not in ("", "합계"):
            in_gyeongnam = False
            current_city = None

        if not in_gyeongnam:
            continue

        # 시군구 갱신
        if sigun and sigun in CITY_CODE_MAP:
            current_city = sigun

        if not current_city or current_city not in CITY_CODE_MAP:
            continue

        # 읍면동 소계 행 (tpg=소계/합계 또는 tpg 없는 읍면동 행)
        is_dong_row = False
        if dong and dong not in SKIP_DONGS:
            if tpg in ("소계", "합계", ""):
                is_dong_row = True

        if not is_dong_row:
            continue

        voters = clean_num(row[4] if len(row) > 4 else 0)
        turnout = clean_num(row[5] if len(row) > 5 else 0)
        votes_list = []
        for k in range(num_cands):
            idx = cand_start_col + k
            votes_list.append(clean_num(row[idx] if len(row) > idx else 0))

        if current_city not in city_dong_results:
            city_dong_results[current_city] = []

        city_dong_results[current_city].append(
            make_dong_result(dong, current_city, voters, turnout, candidates, votes_list)
        )

    for city, dongs in city_dong_results.items():
        print(f"    {city}: {len(dongs)}개 읍면동")

    return city_dong_results


# ===== 지선 파싱 =====

def parse_local_election(filepath, sub_type, label, date):
    """지선 읍면동별 xlsx 파싱"""
    print(f"  파싱: {label} ({sub_type})")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_col=30, values_only=True))
    wb.close()

    # 헤더: 행1=컬럼명, 행2=후보 번호, 행3~=후보 정보(선거구별)
    # 구조: 시도명 | 구시군명 | 선거구명 | 읍면동명 | 구분 | 선거인수 | 투표수 | 후보1 | 후보2 | ...

    city_dong_results = {}
    current_city = None
    current_district = None
    current_candidates = []
    cand_start_col = 7  # 기본값

    for i, row in enumerate(rows):
        r = [str(v or "").strip() for v in row]

        # 첫 행에서 후보 시작 열 찾기
        if i == 0:
            for j, v in enumerate(row):
                if v and "후보" in str(v):
                    cand_start_col = j
                    break
            continue

        sido = r[0]
        sigun = r[1]
        sgg = r[2]  # 선거구명
        dong = r[3]
        gubun = r[4] if len(r) > 4 else ""

        # 경남이 아니면 건너뜀
        if sido and "경상남도" not in sido and sido != "":
            if current_city and current_city in CITY_CODE_MAP:
                pass  # 이미 수집 중
            else:
                continue

        if sido and "경상남도" not in sido and current_city:
            current_city = None  # 다른 시도

        # 후보자 정보 행 (이름\n정당 형태)
        has_cand = sum(1 for c in row[cand_start_col:] if c and "\n" in str(c))
        if has_cand >= 2:
            current_candidates = parse_candidates_from_header(row[cand_start_col:])
            if sigun and sigun in CITY_CODE_MAP:
                current_city = sigun
            elif sgg:
                # 선거구명에서 도시 추출
                for city_name in sorted(CITY_CODE_MAP.keys(), key=len, reverse=True):
                    if sgg.startswith(city_name):
                        current_city = city_name
                        break
            current_district = sgg
            continue

        # 시군구 갱신
        if sigun and sigun in CITY_CODE_MAP:
            current_city = sigun
        if sgg:
            current_district = sgg

        if not current_city or current_city not in CITY_CODE_MAP:
            continue
        if not current_candidates:
            continue

        # 읍면동 소계 행
        is_dong = False
        if dong and dong not in SKIP_DONGS and dong not in ("거소투표", "관외사전투표", "거소·선상투표"):
            if gubun == "소계" or gubun == "":
                is_dong = True

        if not is_dong:
            continue

        voters = clean_num(row[5] if len(row) > 5 else 0)
        turnout = clean_num(row[6] if len(row) > 6 else 0)
        votes_list = []
        for k in range(len(current_candidates)):
            idx = cand_start_col + k
            votes_list.append(clean_num(row[idx] if len(row) > idx else 0))

        if current_city not in city_dong_results:
            city_dong_results[current_city] = []

        city_dong_results[current_city].append(
            make_dong_result(dong, current_district or current_city, voters, turnout, current_candidates, votes_list)
        )

    for city, dongs in city_dong_results.items():
        print(f"    {city}: {len(dongs)}개 읍면동")

    return city_dong_results


# ===== 2018 지선 파싱 (구조가 2022와 다름) =====

def parse_local_2018(filepath, sub_type, label, date):
    """2018 지선 xlsx 파싱
    시장/기초/도의원: 선거종류(0) | 시도(1) | 선거구명(2) | 시도명(3) | 구시군명(4) | 읍면동명(5) | 구분(6) | 선거인수(7) | 투표수(8) | 후보(9~)
    시도지사/교육감/비례: 선거종류(0) | 시도(1) | 선거구명(2) | 구시군명(3) | 읍면동명(4) | 구분(5) | 선거인수(6) | 투표수(7) | 후보(8~)
    """
    print(f"  파싱: {label} ({sub_type})")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_col=30, values_only=True))
    wb.close()

    # 시도 단위 선거는 열 구조가 다름
    is_sido_level = sub_type in ("시도지사", "교육감", "광역비례")
    if is_sido_level:
        COL_SIGUN = 3
        COL_DONG = 4
        COL_GUBUN = 5
        COL_VOTERS = 6
        COL_TURNOUT = 7
        COL_CAND_START = 8
    else:
        COL_SIGUN = 4
        COL_DONG = 5
        COL_GUBUN = 6
        COL_VOTERS = 7
        COL_TURNOUT = 8
        COL_CAND_START = 9

    city_dong_results = {}
    current_city = None
    current_district = None
    current_candidates = []

    for i, row in enumerate(rows):
        r = [str(v or "").strip() for v in row]

        # 행0 = 헤더, 건너뜀
        if i == 0:
            continue

        sido = r[1]
        sgg = r[2]
        sigun = r[COL_SIGUN] if len(r) > COL_SIGUN else ""
        dong = r[COL_DONG] if len(r) > COL_DONG else ""
        gubun = r[COL_GUBUN] if len(r) > COL_GUBUN else ""

        # 후보자 정보 행 감지
        has_cand = sum(1 for c in row[COL_CAND_START:] if c and "\n" in str(c))
        if has_cand >= 2:
            current_candidates = parse_candidates_from_header(row[COL_CAND_START:])
            if sgg:
                current_district = sgg
            if is_sido_level:
                # 시도 단위: 선거구명이 시도명
                current_city = None  # 행마다 구시군에서 갱신
            else:
                for city_name in sorted(CITY_CODE_MAP.keys(), key=len, reverse=True):
                    if sgg.startswith(city_name):
                        current_city = city_name
                        break
            continue

        # 경남 아니면 건너뜀
        if sido and "경상남도" not in sido and sgg and "경상남도" not in sgg:
            current_city = None
            continue

        # 시군구 갱신
        if sigun and sigun in CITY_CODE_MAP:
            current_city = sigun
        elif not is_sido_level and sgg and sgg != current_district:
            current_district = sgg
            for city_name in sorted(CITY_CODE_MAP.keys(), key=len, reverse=True):
                if sgg.startswith(city_name):
                    current_city = city_name
                    break

        if not current_city or current_city not in CITY_CODE_MAP:
            continue
        if not current_candidates:
            continue

        # 읍면동 소계 행
        if dong and dong not in SKIP_DONGS and gubun == "소계":
            voters = clean_num(row[COL_VOTERS] if len(row) > COL_VOTERS else 0)
            turnout = clean_num(row[COL_TURNOUT] if len(row) > COL_TURNOUT else 0)
            votes_list = []
            for k in range(len(current_candidates)):
                idx = COL_CAND_START + k
                votes_list.append(clean_num(row[idx] if len(row) > idx else 0))

            if sum(votes_list) == 0:
                continue

            if current_city not in city_dong_results:
                city_dong_results[current_city] = []

            district = current_district if not is_sido_level else current_city
            city_dong_results[current_city].append(
                make_dong_result(dong, district, voters, turnout, current_candidates, votes_list)
            )

    for city, dongs in city_dong_results.items():
        print(f"    {city}: {len(dongs)}개 읍면동")

    return city_dong_results


# ===== 총선 파싱 =====

def parse_assembly_2024(filepath, label, date):
    """2024 총선 xlsx 파싱"""
    print(f"  파싱: {label}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_col=30, values_only=True))
    wb.close()

    # 구조: 시도명 | 선거구명 | 구시군명 | 읍면동명 | 투표타입 | 선거인수 | 투표수 | 후보...
    city_dong_results = {}
    current_city = None
    current_district = None
    current_candidates = []
    cand_start_col = 7

    in_gyeongnam = False

    for i, row in enumerate(rows):
        r = [str(v or "").strip() for v in row]

        # 헤더 행 건너뛰기 (제목/컬럼명 행만)
        if r[0] in ("시도명", "개표단위별 개표결과", "[국회의원선거][전체]"):
            if r[0] == "시도명":
                for j, v in enumerate(row):
                    if v and "후보" in str(v):
                        cand_start_col = j
                        break
            continue
        # 완전 빈 행
        if all(v is None or str(v).strip() == "" for v in row[:10]):
            continue

        sido = r[0]
        sgg = r[1]   # 선거구명
        sigun = r[2]  # 구시군명
        dong = r[3]
        vote_type = r[4] if len(r) > 4 else ""

        # 시도 전환 감지
        if sido and "경상남도" in sido:
            in_gyeongnam = True
        elif sido and sido not in ("", "합계"):
            in_gyeongnam = False
            current_city = None

        # 후보자 정보 행
        has_cand = sum(1 for c in row[cand_start_col:] if c and "\n" in str(c))
        if has_cand >= 2:
            current_candidates = parse_candidates_from_header(row[cand_start_col:])
            if in_gyeongnam and sigun and sigun in CITY_CODE_MAP:
                current_city = sigun
            current_district = sgg
            continue

        if not in_gyeongnam:
            continue

        if sigun and sigun in CITY_CODE_MAP:
            current_city = sigun
        if sgg:
            current_district = sgg

        if not current_city or current_city not in CITY_CODE_MAP:
            continue
        if not current_candidates:
            continue

        # 읍면동 행 (dong이 있고 vote_type이 비어있거나 소계)
        if dong and dong not in SKIP_DONGS and dong not in ("거소·선상투표", "관외사전투표", "재외투표", "국외부재자투표"):
            if vote_type in ("", "소계"):
                voters = clean_num(row[5] if len(row) > 5 else 0)
                turnout = clean_num(row[6] if len(row) > 6 else 0)
                votes_list = []
                for k in range(len(current_candidates)):
                    idx = cand_start_col + k
                    votes_list.append(clean_num(row[idx] if len(row) > idx else 0))

                if current_city not in city_dong_results:
                    city_dong_results[current_city] = []
                city_dong_results[current_city].append(
                    make_dong_result(dong, current_district or current_city, voters, turnout, current_candidates, votes_list)
                )

    for city, dongs in city_dong_results.items():
        print(f"    {city}: {len(dongs)}개 읍면동")
    return city_dong_results


def parse_assembly_2020(base_dir, label, date):
    """2020 총선 — 시군구별 개별 파일 파싱"""
    print(f"  파싱: {label}")
    gyeongnam_dir = os.path.join(base_dir, "지역구", "16경남")
    if not os.path.exists(gyeongnam_dir):
        print(f"    경남 폴더 없음: {gyeongnam_dir}")
        return {}

    city_dong_results = {}
    for filename in sorted(os.listdir(gyeongnam_dir)):
        if not filename.endswith(".xlsx"):
            continue
        filepath = os.path.join(gyeongnam_dir, filename)
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(max_col=20, values_only=True))
        wb.close()

        # 파일명에서 도시 추출
        fname = filename.replace("개표상황(투표구별)_", "").replace(".xlsx", "")
        # "밀양시의령군함안군창녕군_밀양시" → 마지막 _뒤가 실제 시군구
        if "_" in fname:
            city_name = fname.split("_")[-1]
        else:
            # "진주시갑" → "진주시"
            city_name = fname
            for suffix in ("갑", "을", "병", "정"):
                if city_name.endswith(suffix):
                    city_name = city_name[:-1]
                    break

        if city_name not in CITY_CODE_MAP:
            continue

        # 구조: 행1=제목, 행2=헤더, 행3=후보자, 행4~=데이터
        candidates = []
        cand_start = 5  # 기본

        for i, row in enumerate(rows[:5]):
            cands = parse_candidates_from_header(row)
            if len(cands) >= 2:
                candidates = cands
                for j, c in enumerate(row):
                    if c and "\n" in str(c):
                        cand_start = j
                        break
                break

        if not candidates:
            continue

        district = fname.split("_")[0] if "_" in fname else fname

        for row in rows:
            r = [str(v or "").strip() for v in row]
            # 읍면동 행 찾기
            dong = None
            for j in range(4):
                if len(r) > j and r[j] and r[j] not in SKIP_DONGS and ("동" in r[j] or "면" in r[j] or "읍" in r[j]):
                    # 소계 행인지 확인
                    dong_candidate = r[j]
                    # 투표구명이 없는 소계 행
                    next_col = r[j+1] if j+1 < len(r) else ""
                    if next_col in ("소계", ""):
                        dong = dong_candidate
                        break

            if not dong:
                continue

            # voters/turnout 열 위치 추정 (후보 시작 -2, -1)
            voters = clean_num(row[cand_start - 2] if cand_start >= 2 else 0)
            turnout = clean_num(row[cand_start - 1] if cand_start >= 1 else 0)
            votes_list = [clean_num(row[cand_start + k] if cand_start + k < len(row) else 0) for k in range(len(candidates))]

            if sum(votes_list) == 0:
                continue

            if city_name not in city_dong_results:
                city_dong_results[city_name] = []
            city_dong_results[city_name].append(
                make_dong_result(dong, district, voters, turnout, candidates, votes_list)
            )

    for city, dongs in city_dong_results.items():
        print(f"    {city}: {len(dongs)}개 읍면동")
    return city_dong_results


# ===== 메인 =====

def merge_dong_results(elections_file, dong_results_by_label):
    """기존 elections JSON에 dongResults 병합"""
    with open(elections_file, "r", encoding="utf-8") as f:
        elections = json.load(f)

    for entry in elections:
        label = entry["label"]
        if label in dong_results_by_label:
            entry["dongResults"] = dong_results_by_label[label]

    with open(elections_file, "w", encoding="utf-8") as f:
        json.dump(elections, f, ensure_ascii=False, indent=2)


def main():
    # 대선 파싱
    presidential_files = [
        (os.path.join(NEC_DIR, "제21대_대통령선거_개표결과.xlsx"), "2025 대선 (제21대)", "2025.06.03"),
        (os.path.join(NEC_DIR, "2022개표단위별_개표결과_대통령선거_전체.xlsx"), "2022 대선 (제20대)", "2022.03.09"),
        (os.path.join(NEC_DIR, "제19대 대통령선거 개표자료.xlsx"), "2017 대선 (제19대)", "2017.05.09"),
    ]

    # label → {city: dongResults}
    all_results = {}

    for filepath, label, date in presidential_files:
        if not os.path.exists(filepath):
            print(f"  파일 없음: {filepath}")
            continue
        results = parse_presidential(filepath, label, date)
        all_results[label] = results

    # 2024 총선
    f2024 = os.path.join(EXTRACT_DIR, "2024총선개표결과", "1. 개표단위별 개표결과(지역구) -전국.xlsx")
    if os.path.exists(f2024):
        results = parse_assembly_2024(f2024, "2024 총선 (제22대)", "2024.04.10")
        all_results["2024 총선 (제22대)"] = results

    # 2020 총선
    f2020 = os.path.join(EXTRACT_DIR, "2020제21대 국회의원선거(재보궐 포함) 투표구별 개표결과")
    if os.path.exists(f2020):
        results = parse_assembly_2020(f2020, "2020 총선 (제21대)", "2020.04.15")
        all_results["2020 총선 (제21대)"] = results

    # 제8회 지선 (2022) 읍면동별
    local8_dir = os.path.join(EXTRACT_DIR, "제8회_전국동시지방선거_읍면동별_개표결과-게시판게시")
    local8_files = [
        ("구시군장선거.xlsx", "시장", "2022 지선 (제8회) 시장"),
        ("시도지사선거.xlsx", "시도지사", "2022 지선 (제8회) 시도지사"),
        ("구시군의원선거.xlsx", "기초의원", "2022 지선 (제8회) 기초의원"),
        ("시도의원선거.xlsx", "도의원", "2022 지선 (제8회) 도의원"),
        ("교육감선거.xlsx", "교육감", "2022 지선 (제8회) 교육감"),
        ("기초비례의원선거.xlsx", "기초비례", "2022 지선 (제8회) 기초비례"),
        ("광역비례의원선거.xlsx", "광역비례", "2022 지선 (제8회) 광역비례"),
    ]

    all_local_results = {}
    for fname, sub_type, label in local8_files:
        filepath = os.path.join(local8_dir, fname)
        if os.path.exists(filepath):
            results = parse_local_election(filepath, sub_type, label, "2022.06.01")
            all_local_results[label] = results

    # 제7회 지선 (2018) 읍면동별 — 2018은 구조가 다름
    local7_dir = os.path.join(EXTRACT_DIR, "2018전국동시지방선거 개표결과(제7회)")
    local7_files = [
        ("20180619-7지선-02-(구시군의장)_읍면동별개표자료.xlsx", "시장", "2018 지선 (제7회) 시장"),
        ("20180619-7지선-01-(시도지사)_읍면동별개표자료.xlsx", "시도지사", "2018 지선 (제7회) 시도지사"),
        ("20180619-7지선-04-(구시군의회의원)_읍면동별개표자료.xlsx", "기초의원", "2018 지선 (제7회) 기초의원"),
        ("20180619-7지선-03-(시도의회의원)_읍면동별개표자료.xlsx", "도의원", "2018 지선 (제7회) 도의원"),
        ("20180619-7지선-07-(교육감)_읍면동별개표자료.xlsx", "교육감", "2018 지선 (제7회) 교육감"),
        ("20180619-7지선-06-(기초비례)_읍면동별개표자료.xlsx", "기초비례", "2018 지선 (제7회) 기초비례"),
        ("20180619-7지선-05-(광역비례)_읍면동별개표자료.xlsx", "광역비례", "2018 지선 (제7회) 광역비례"),
    ]

    for fname, sub_type, label in local7_files:
        filepath = os.path.join(local7_dir, fname)
        if os.path.exists(filepath):
            results = parse_local_2018(filepath, sub_type, label, "2018.06.13")
            all_local_results[label] = results

    # 도시별 JSON에 dongResults 병합
    print(f"\n{'='*50}")
    print("dongResults 병합 중...")

    for city_name, city_code in CITY_CODE_MAP.items():
        el_file = os.path.join(OUTPUT_DIR, f"{city_code}-elections.json")
        lo_file = os.path.join(OUTPUT_DIR, f"{city_code}-local-elections.json")

        # 일반 선거 (대선/총선) dongResults
        if os.path.exists(el_file):
            city_el_dongs = {}
            for label, city_results in all_results.items():
                if city_name in city_results:
                    city_el_dongs[label] = city_results[city_name]
            if city_el_dongs:
                merge_dong_results(el_file, city_el_dongs)
                labels_with = [l for l in city_el_dongs]
                print(f"  [{city_code}] {city_name} 일반: {', '.join(labels_with)}")

        # 지선 dongResults
        if os.path.exists(lo_file):
            city_lo_dongs = {}
            for label, city_results in all_local_results.items():
                if city_name in city_results:
                    city_lo_dongs[label] = city_results[city_name]
            if city_lo_dongs:
                merge_dong_results(lo_file, city_lo_dongs)
                count = sum(len(v) for v in city_lo_dongs.values())
                print(f"  [{city_code}] {city_name} 지선: {len(city_lo_dongs)}종 {count}개 읍면동")

    print("\n완료!")


if __name__ == "__main__":
    main()
