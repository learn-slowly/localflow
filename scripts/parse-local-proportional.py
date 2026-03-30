#!/usr/bin/env python3
"""2018/2022 지선 광역비례·기초비례 읍면동별 결과 → local-elections JSON에 추가"""

import json
import os
import openpyxl

BASE = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE, "..", "public", "data", "elections")

SOURCES = [
    {
        "date": "2022.06.01",
        "entries": [
            {
                "subType": "광역비례",
                "path": "src/data/nec/제8회_전국동시지방선거_읍면동별_개표결과-게시판게시/광역비례의원선거.xlsx",
                "format": "2022",
            },
            {
                "subType": "기초비례",
                "path": "src/data/nec/제8회_전국동시지방선거_읍면동별_개표결과-게시판게시/기초비례의원선거.xlsx",
                "format": "2022",
            },
        ],
    },
    {
        "date": "2018.06.13",
        "entries": [
            {
                "subType": "광역비례",
                "path": "src/data/nec/2018전국동시지방선거 개표결과(제7회)/20180619-7지선-05-(광역비례)_읍면동별개표자료.xlsx",
            },
            {
                "subType": "기초비례",
                "path": "src/data/nec/2018전국동시지방선거 개표결과(제7회)/20180619-7지선-06-(기초비례)_읍면동별개표자료.xlsx",
            },
        ],
    },
]


def parse_2022_xlsx(path, city_name="진주시"):
    """2022 지선 비례 xlsx 파싱
    광역비례: [시도명, 구시군명, 읍면동명, 구분, 선거인수, 투표수, 정당별...]
    기초비례: [시도명, 구시군, 선거구, 읍면동명, 구분, 선거인수, 투표수, 정당별...]
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))

    # 기초비례 vs 광역비례 구분 — row1의 col2가 "선거구"
    header = list(all_rows[0]) if all_rows else []
    is_basic = len(header) > 2 and "선거구" in str(header[2] or "")

    if is_basic:
        sgg_col, emd_col, tpg_col, voters_col, turnout_col, party_start = 1, 3, 4, 5, 6, 7
    else:
        sgg_col, emd_col, tpg_col, voters_col, turnout_col, party_start = 1, 2, 3, 4, 5, 6

    # 진주시 정당 헤더 찾기
    parties = []
    for row in all_rows:
        vals = list(row)
        if len(vals) > sgg_col and city_name in str(vals[sgg_col] or ""):
            if len(vals) > party_start and vals[party_start]:
                raw = [str(v).strip() for v in vals[party_start:] if v and str(v).strip()]
                parties = [p for p in raw if p not in ("계", "")]
                break

    if not parties:
        print(f"  정당 헤더 못 찾음")
        return None, []

    print(f"  정당 {len(parties)}개: {parties[:5]}...")

    SKIP = {"합계", "거소투표", "관외사전투표", "재외투표"}
    dong_results = []
    in_city = False

    for row in all_rows:
        vals = list(row)
        if len(vals) <= party_start:
            continue

        sgg = str(vals[sgg_col] or "").strip()
        emd = str(vals[emd_col] or "").strip()
        tpg = str(vals[tpg_col] or "").strip()

        if city_name in sgg:
            in_city = True
        elif in_city and sgg and city_name not in sgg:
            break

        if not in_city or emd in SKIP or not emd or tpg != "소계":
            continue

        voters = parse_int(vals[voters_col])
        turnout = parse_int(vals[turnout_col])
        party_votes = {}
        valid = 0
        for j, party in enumerate(parties):
            idx = party_start + j
            v = parse_int(vals[idx]) if idx < len(vals) else 0
            if v > 0:
                party_votes[party] = v
                valid += v

        dong_results.append({
            "dong": emd,
            "voters": voters,
            "turnout": turnout,
            "valid": valid,
            "parties": party_votes,
        })

    # 합산
    total = {"voters": 0, "turnout": 0, "valid": 0, "invalid": 0, "parties": {}}
    for d in dong_results:
        total["voters"] += d["voters"]
        total["turnout"] += d["turnout"]
        total["valid"] += d["valid"]
        for party, votes in d["parties"].items():
            total["parties"][party] = total["parties"].get(party, 0) + votes

    total["parties"] = [{"party": p, "votes": v} for p, v in sorted(total["parties"].items(), key=lambda x: -x[1])]
    return total, dong_results


def parse_int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    try:
        return int(float(s))
    except:
        return 0


def parse_2018_xlsx(path, city_name="진주시"):
    """2018 지선 비례 xlsx 파싱
    광역비례: [선거종류, 시도, 시도명, 구시군명, 읍면동명, 구분, 선거인수, 투표수, 정당별...]
    기초비례: [선거종류, 시도, 선거구명, 시도명, 구시군명, 읍면동명, 구분, 선거인수, 투표수, 정당별...]
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))

    # 기초비례 vs 광역비례 구분 — row1의 col2가 "선거구명"이면 기초비례
    header = list(all_rows[0]) if all_rows else []
    is_basic = len(header) > 2 and str(header[2] or "").strip() == "선거구명"

    if is_basic:
        # 기초비례: sgg=col4, emd=col5, tpg=col6, voters=col7, turnout=col8, parties from col9
        sgg_col, emd_col, tpg_col, voters_col, turnout_col, party_start = 4, 5, 6, 7, 8, 9
    else:
        # 광역비례: sgg=col3, emd=col4, tpg=col5, voters=col6, turnout=col7, parties from col8
        sgg_col, emd_col, tpg_col, voters_col, turnout_col, party_start = 3, 4, 5, 6, 7, 8

    # 정당 이름 찾기 — 도시가 포함된 헤더 행에서 추출
    parties = []
    for row in all_rows:
        vals = list(row)
        if len(vals) > party_start and vals[party_start]:
            v = str(vals[party_start]).strip()
            # 진주시가 포함된 행의 정당 헤더, 또는 첫 번째 정당 헤더
            if is_basic:
                if city_name in str(vals[2] or "") and any(
                    kw in v for kw in ["더불어", "자유", "바른", "민주", "국민", "정의"]
                ):
                    raw = [str(x).strip() for x in vals[party_start:] if x and str(x).strip()]
                    parties = [p for p in raw if p != "계"]
                    break
            else:
                if any(kw in v for kw in ["더불어", "자유", "바른", "민주", "국민", "정의"]):
                    raw = [str(x).strip() for x in vals[party_start:] if x and str(x).strip()]
                    parties = [p for p in raw if p != "계"]
                    break

    if not parties:
        print(f"  정당 헤더 못 찾음")
        return None, []

    print(f"  정당 {len(parties)}개: {parties[:5]}...")

    SKIP = {"계", "거소투표", "관외사전투표", "재외투표", "합계"}
    dong_results = []
    in_city = False

    for row in all_rows:
        vals = list(row)
        if len(vals) <= party_start:
            continue

        sgg = str(vals[sgg_col] or "").strip()
        emd = str(vals[emd_col] or "").strip()
        tpg = str(vals[tpg_col] or "").strip()

        if city_name in sgg:
            in_city = True
        elif in_city and sgg and city_name not in sgg:
            break

        if not in_city:
            continue

        if emd in SKIP or not emd or tpg != "소계":
            continue

        voters = parse_int(vals[voters_col])
        turnout = parse_int(vals[turnout_col])
        party_votes = {}
        valid = 0
        for j, party in enumerate(parties):
            idx = party_start + j
            v = parse_int(vals[idx]) if idx < len(vals) else 0
            if v > 0:
                party_votes[party] = v
                valid += v

        dong_results.append({
            "dong": emd,
            "voters": voters,
            "turnout": turnout,
            "valid": valid,
            "parties": party_votes,
        })

    # 합산
    total = {"voters": 0, "turnout": 0, "valid": 0, "invalid": 0, "parties": {}}
    for d in dong_results:
        total["voters"] += d["voters"]
        total["turnout"] += d["turnout"]
        total["valid"] += d["valid"]
        for party, votes in d["parties"].items():
            total["parties"][party] = total["parties"].get(party, 0) + votes

    total["parties"] = [{"party": p, "votes": v} for p, v in sorted(total["parties"].items(), key=lambda x: -x[1])]
    return total, dong_results


def main():
    json_path = os.path.join(DATA_DIR, "48170-local-elections.json")
    with open(json_path, "r", encoding="utf-8") as f:
        elections = json.load(f)

    for src in SOURCES:
        date = src["date"]
        for entry in src["entries"]:
            sub_type = entry["subType"]
            path = os.path.join(BASE, "..", entry["path"])

            print(f"\n=== {date} {sub_type} ===")
            if not os.path.exists(path):
                print(f"  파일 없음: {path}")
                continue

            if entry.get("format") == "2022":
                total, dong_results = parse_2022_xlsx(path)
            else:
                total, dong_results = parse_2018_xlsx(path)
            if not dong_results:
                print("  데이터 없음")
                continue

            print(f"  {len(dong_results)}개 읍면동")
            top = sorted(total["parties"], key=lambda p: -p["votes"])[:5]
            for p in top:
                pct = (p["votes"] / total["valid"] * 100) if total["valid"] else 0
                print(f"    {p['party']}: {p['votes']:,}표 ({pct:.1f}%)")

            # 기존 데이터에서 해당 subType 찾아서 추가
            for elec in elections:
                if elec.get("date") == date and elec.get("subType") == sub_type:
                    elec["proportional"] = total
                    elec["proportionalByDong"] = dong_results
                    print(f"  → '{elec['label']}'에 추가 완료")
                    break
            else:
                print(f"  매칭 실패: date={date}, subType={sub_type}")

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(elections, f, ensure_ascii=False, indent=2)
    print(f"\n저장: {json_path}")


if __name__ == "__main__":
    main()
