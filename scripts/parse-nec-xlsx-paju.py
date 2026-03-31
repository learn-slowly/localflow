#!/usr/bin/env python3
"""NEC(선관위) xlsx 파일에서 파주시 읍면동별 선거결과를 추출하여
elections/local-elections JSON에 dongResults를 추가

기존 parse-nec-xlsx.py를 파주시(경기도)에 맞게 조정한 버전
"""

import json
import os
import openpyxl

NEC_DIR = os.path.join(os.path.dirname(__file__), "..", "src", "data", "nec")
EXTRACT_DIR = os.path.join(os.path.dirname(__file__), "..", "src", "data", "nec")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "public", "data", "elections")

CITY_CODE = "41480"
CITY_NAME = "파주시"
SIDO_NAME = "경기도"

SKIP_DONGS = {"합계", "거소·선상투표", "거소투표", "관외사전투표", "재외투표", "", "계", "국외부재자투표"}


def clean_num(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    try:
        return int(float(s))
    except:
        return 0


def parse_candidates_from_header(row):
    candidates = []
    for cell in row:
        if cell and "\n" in str(cell):
            text = str(cell).replace("_x000D_", "").replace("\r", "")
            parts = text.split("\n")
            party = parts[0].strip()
            name = parts[1].strip() if len(parts) > 1 else ""
            if name:
                candidates.append({"name": name, "party": party})
    return candidates


def make_dong_result(dong_name, district, voters, turnout, candidates, votes_list):
    valid = sum(votes_list)
    rates = {}
    votes_map = {}
    for cand, v in zip(candidates, votes_list):
        if valid > 0:
            rates[cand["name"]] = round(v / valid * 100, 1)
        votes_map[cand["name"]] = v
    return {
        "dong": dong_name,
        "district": district,
        "voters": voters,
        "turnout": turnout,
        "rates": rates,
        "votes": votes_map,
    }


# ===== 대선 파싱 =====

def parse_presidential(filepath, label, date):
    print(f"  파싱: {label}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_col=30, values_only=True))
    wb.close()

    candidates = []
    header_row = None
    cand_start_col = None

    # 후보자 행 찾기
    for i, row in enumerate(rows[:5]):
        cands = parse_candidates_from_header(row)
        if len(cands) >= 3:
            candidates = cands
            header_row = i
            for j, c in enumerate(row):
                if c and "\n" in str(c):
                    cand_start_col = j
                    break
            break

    if not candidates:
        for i, row in enumerate(rows):
            r = [str(v or "") for v in row]
            if "시도" in r[0] or "시도명" in r[0]:
                if i + 1 < len(rows):
                    next_row = rows[i + 1]
                    cands = parse_candidates_from_header(next_row)
                    if len(cands) >= 3:
                        candidates = cands
                        header_row = i + 1
                        for j, c in enumerate(next_row):
                            if c and "\n" in str(c):
                                cand_start_col = j
                                break
                        break

    if not candidates:
        print(f"    후보자 정보를 찾을 수 없음")
        return []

    print(f"    후보자 {len(candidates)}명: {', '.join(c['name'] for c in candidates)}")

    dong_results = []
    current_city = None
    data_start = header_row + 1 if header_row is not None else 2
    num_cands = len(candidates)
    in_gyeonggi = False

    for row in rows[data_start:]:
        r = [str(v or "").strip() for v in row]

        sido = r[0] if r[0] else None
        sigun = r[1] if len(r) > 1 and r[1] else None
        dong = r[2] if len(r) > 2 and r[2] else None
        tpg = r[3] if len(r) > 3 and r[3] else None

        if sido and SIDO_NAME in sido:
            in_gyeonggi = True
        elif sido and sido not in ("", "합계"):
            in_gyeonggi = False
            current_city = None

        if not in_gyeonggi:
            continue

        if sigun and CITY_NAME in sigun:
            current_city = CITY_NAME
        elif sigun and sigun != CITY_NAME:
            current_city = sigun

        if current_city != CITY_NAME:
            continue

        if dong and dong not in SKIP_DONGS:
            if tpg in ("소계", "합계", ""):
                voters = clean_num(row[4] if len(row) > 4 else 0)
                turnout = clean_num(row[5] if len(row) > 5 else 0)
                votes_list = []
                for k in range(num_cands):
                    idx = cand_start_col + k
                    votes_list.append(clean_num(row[idx] if len(row) > idx else 0))

                if sum(votes_list) > 0:
                    dong_results.append(
                        make_dong_result(dong, CITY_NAME, voters, turnout, candidates, votes_list)
                    )

    print(f"    파주시: {len(dong_results)}개 읍면동")
    return dong_results


# ===== 총선 파싱 =====

def parse_assembly_2024(filepath, label, date):
    print(f"  파싱: {label}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_col=30, values_only=True))
    wb.close()

    dong_results = []
    current_city = None
    current_district = None
    current_candidates = []
    cand_start_col = 7
    in_gyeonggi = False

    for i, row in enumerate(rows):
        r = [str(v or "").strip() for v in row]

        if r[0] in ("시도명", "개표단위별 개표결과", "[국회의원선거][전체]"):
            if r[0] == "시도명":
                for j, v in enumerate(row):
                    if v and "후보" in str(v):
                        cand_start_col = j
                        break
            continue

        if all(v is None or str(v).strip() == "" for v in row[:10]):
            continue

        sido = r[0]
        sgg = r[1]
        sigun = r[2]
        dong = r[3]
        vote_type = r[4] if len(r) > 4 else ""

        if sido and SIDO_NAME in sido:
            in_gyeonggi = True
        elif sido and sido not in ("", "합계"):
            in_gyeonggi = False
            current_city = None

        has_cand = sum(1 for c in row[cand_start_col:] if c and "\n" in str(c))
        if has_cand >= 2:
            current_candidates = parse_candidates_from_header(row[cand_start_col:])
            current_district = sgg
            # 총선은 선거구명에 도시명 포함 (예: "파주시갑", "파주시을")
            if in_gyeonggi and sgg and CITY_NAME in sgg:
                current_city = CITY_NAME
            elif in_gyeonggi and sigun == CITY_NAME:
                current_city = CITY_NAME
            else:
                current_city = None  # 다른 도시 선거구
            continue

        if not in_gyeonggi:
            continue

        if sigun == CITY_NAME:
            current_city = CITY_NAME
        elif sigun and sigun != CITY_NAME and CITY_NAME not in (current_district or ""):
            current_city = None
        if sgg:
            current_district = sgg
            if CITY_NAME in sgg:
                current_city = CITY_NAME
            elif current_city == CITY_NAME and CITY_NAME not in sgg and sgg not in ("", "합계"):
                current_city = None

        if current_city != CITY_NAME or not current_candidates:
            continue

        if dong and dong not in SKIP_DONGS and dong not in ("거소·선상투표", "관외사전투표", "재외투표"):
            if vote_type in ("", "소계"):
                voters = clean_num(row[5] if len(row) > 5 else 0)
                turnout = clean_num(row[6] if len(row) > 6 else 0)
                votes_list = []
                for k in range(len(current_candidates)):
                    idx = cand_start_col + k
                    votes_list.append(clean_num(row[idx] if len(row) > idx else 0))
                if sum(votes_list) > 0:
                    dong_results.append(
                        make_dong_result(dong, current_district or CITY_NAME, voters, turnout, current_candidates, votes_list)
                    )

    print(f"    파주시: {len(dong_results)}개 읍면동")
    return dong_results


def parse_assembly_2020(base_dir, label, date):
    print(f"  파싱: {label}")
    gyeonggi_dir = os.path.join(base_dir, "지역구", "9경기")
    if not os.path.exists(gyeonggi_dir):
        for d in os.listdir(os.path.join(base_dir, "지역구")):
            if "경기" in d:
                gyeonggi_dir = os.path.join(base_dir, "지역구", d)
                break
    if not os.path.exists(gyeonggi_dir):
        print(f"    경기 폴더 없음: {gyeonggi_dir}")
        return []

    dong_results = []
    for filename in sorted(os.listdir(gyeonggi_dir)):
        if not filename.endswith(".xlsx"):
            continue
        # 파주시 관련 파일만 (파주시갑, 파주시을 등)
        if "파주시" not in filename:
            continue

        filepath = os.path.join(gyeonggi_dir, filename)
        print(f"    파일: {filename}")
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(max_col=20, values_only=True))
        wb.close()

        fname = filename.replace("개표상황(투표구별)_", "").replace(".xlsx", "")
        district = fname.split("_")[0] if "_" in fname else fname

        candidates = []
        cand_start = 5

        for i, row in enumerate(rows[:5]):
            cands = parse_candidates_from_header(row)
            if len(cands) >= 2:
                candidates = cands
                for j, c in enumerate(row):
                    if c and "\n" in str(c):
                        cand_start = j
                        break
                break

        if not candidates:
            continue

        for row in rows:
            r = [str(v or "").strip() for v in row]
            dong = None
            for j in range(4):
                if len(r) > j and r[j] and r[j] not in SKIP_DONGS and ("동" in r[j] or "면" in r[j] or "읍" in r[j]):
                    next_col = r[j+1] if j+1 < len(r) else ""
                    if next_col in ("소계", ""):
                        dong = r[j]
                        break

            if not dong:
                continue

            voters = clean_num(row[cand_start - 2] if cand_start >= 2 else 0)
            turnout = clean_num(row[cand_start - 1] if cand_start >= 1 else 0)
            votes_list = [clean_num(row[cand_start + k] if cand_start + k < len(row) else 0) for k in range(len(candidates))]

            if sum(votes_list) == 0:
                continue

            dong_results.append(
                make_dong_result(dong, district, voters, turnout, candidates, votes_list)
            )

    print(f"    파주시: {len(dong_results)}개 읍면동")
    return dong_results


# ===== 지선 파싱 =====

def parse_local_election(filepath, sub_type, label, date):
    """제8회 지선 xlsx 파싱"""
    print(f"  파싱: {label} ({sub_type})")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_col=30, values_only=True))
    wb.close()

    dong_results = []
    current_city = None
    current_district = None
    current_candidates = []
    cand_start_col = 7

    for i, row in enumerate(rows):
        r = [str(v or "").strip() for v in row]

        if i == 0:
            for j, v in enumerate(row):
                if v and "후보" in str(v):
                    cand_start_col = j
                    break
            continue

        sido = r[0]
        sigun = r[1]
        sgg = r[2]
        dong = r[3]
        gubun = r[4] if len(r) > 4 else ""

        if sido and SIDO_NAME not in sido and sido != "":
            if current_city == CITY_NAME:
                pass
            else:
                continue

        if sido and SIDO_NAME not in sido and current_city:
            current_city = None

        has_cand = sum(1 for c in row[cand_start_col:] if c and "\n" in str(c))
        if has_cand >= 2:
            current_candidates = parse_candidates_from_header(row[cand_start_col:])
            if sigun == CITY_NAME:
                current_city = CITY_NAME
            elif sgg and sgg.startswith(CITY_NAME):
                current_city = CITY_NAME
            else:
                current_city = sigun
            current_district = sgg
            continue

        if sigun == CITY_NAME:
            current_city = CITY_NAME
        elif sigun and sigun != CITY_NAME:
            current_city = sigun
        if sgg:
            current_district = sgg

        if current_city != CITY_NAME or not current_candidates:
            continue

        if dong and dong not in SKIP_DONGS and dong not in ("거소투표", "관외사전투표", "거소·선상투표"):
            if gubun == "소계" or gubun == "":
                voters = clean_num(row[5] if len(row) > 5 else 0)
                turnout = clean_num(row[6] if len(row) > 6 else 0)
                votes_list = []
                for k in range(len(current_candidates)):
                    idx = cand_start_col + k
                    votes_list.append(clean_num(row[idx] if len(row) > idx else 0))

                if sum(votes_list) > 0:
                    dong_results.append(
                        make_dong_result(dong, current_district or CITY_NAME, voters, turnout, current_candidates, votes_list)
                    )

    print(f"    파주시: {len(dong_results)}개 읍면동")
    return dong_results


def parse_local_2018(filepath, sub_type, label, date):
    """제7회 지선 xlsx 파싱"""
    print(f"  파싱: {label} ({sub_type})")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_col=30, values_only=True))
    wb.close()

    is_sido_level = sub_type in ("시도지사", "교육감", "광역비례")
    if is_sido_level:
        COL_SIGUN = 3
        COL_DONG = 4
        COL_GUBUN = 5
        COL_VOTERS = 6
        COL_TURNOUT = 7
        COL_CAND_START = 8
    else:
        COL_SIGUN = 4
        COL_DONG = 5
        COL_GUBUN = 6
        COL_VOTERS = 7
        COL_TURNOUT = 8
        COL_CAND_START = 9

    dong_results = []
    current_city = None
    current_district = None
    current_candidates = []

    for i, row in enumerate(rows):
        r = [str(v or "").strip() for v in row]

        if i == 0:
            continue

        sido = r[1]
        sgg = r[2]
        sigun = r[COL_SIGUN] if len(r) > COL_SIGUN else ""
        dong = r[COL_DONG] if len(r) > COL_DONG else ""
        gubun = r[COL_GUBUN] if len(r) > COL_GUBUN else ""

        has_cand = sum(1 for c in row[COL_CAND_START:] if c and "\n" in str(c))
        if has_cand >= 2:
            current_candidates = parse_candidates_from_header(row[COL_CAND_START:])
            if sgg:
                current_district = sgg
            if is_sido_level:
                current_city = None
            else:
                if sgg and sgg.startswith(CITY_NAME):
                    current_city = CITY_NAME
                else:
                    current_city = None
            continue

        if sido and SIDO_NAME not in sido and sgg and SIDO_NAME not in sgg:
            current_city = None
            continue

        if sigun == CITY_NAME:
            current_city = CITY_NAME
        elif sigun and sigun != CITY_NAME:
            current_city = sigun
        elif not is_sido_level and sgg and sgg != current_district:
            current_district = sgg
            if sgg.startswith(CITY_NAME):
                current_city = CITY_NAME
            else:
                current_city = None

        if current_city != CITY_NAME or not current_candidates:
            continue

        if dong and dong not in SKIP_DONGS and gubun == "소계":
            voters = clean_num(row[COL_VOTERS] if len(row) > COL_VOTERS else 0)
            turnout = clean_num(row[COL_TURNOUT] if len(row) > COL_TURNOUT else 0)
            votes_list = []
            for k in range(len(current_candidates)):
                idx = COL_CAND_START + k
                votes_list.append(clean_num(row[idx] if len(row) > idx else 0))

            if sum(votes_list) == 0:
                continue

            district = current_district if not is_sido_level else CITY_NAME
            dong_results.append(
                make_dong_result(dong, district, voters, turnout, current_candidates, votes_list)
            )

    print(f"    파주시: {len(dong_results)}개 읍면동")
    return dong_results


# ===== 메인 =====

def merge_dong_results(elections_file, dong_results_by_label):
    with open(elections_file, "r", encoding="utf-8") as f:
        elections = json.load(f)

    for entry in elections:
        label = entry["label"]
        if label in dong_results_by_label:
            entry["dongResults"] = dong_results_by_label[label]

    with open(elections_file, "w", encoding="utf-8") as f:
        json.dump(elections, f, ensure_ascii=False, indent=2)


def main():
    # 대선
    presidential_files = [
        (os.path.join(NEC_DIR, "제21대_대통령선거_개표결과.xlsx"), "2025 대선 (제21대)", "2025.06.03"),
        (os.path.join(NEC_DIR, "2022개표단위별_개표결과_대통령선거_전체.xlsx"), "2022 대선 (제20대)", "2022.03.09"),
        (os.path.join(NEC_DIR, "제19대 대통령선거 개표자료.xlsx"), "2017 대선 (제19대)", "2017.05.09"),
    ]

    el_dongs = {}
    for filepath, label, date in presidential_files:
        if not os.path.exists(filepath):
            print(f"  파일 없음: {filepath}")
            continue
        results = parse_presidential(filepath, label, date)
        if results:
            el_dongs[label] = results

    # 2024 총선
    f2024 = os.path.join(EXTRACT_DIR, "2024총선개표결과", "1. 개표단위별 개표결과(지역구) -전국.xlsx")
    if os.path.exists(f2024):
        results = parse_assembly_2024(f2024, "2024 총선 (제22대)", "2024.04.10")
        if results:
            el_dongs["2024 총선 (제22대)"] = results

    # 2020 총선
    f2020 = os.path.join(EXTRACT_DIR, "2020제21대 국회의원선거(재보궐 포함) 투표구별 개표결과")
    if os.path.exists(f2020):
        results = parse_assembly_2020(f2020, "2020 총선 (제21대)", "2020.04.15")
        if results:
            el_dongs["2020 총선 (제21대)"] = results

    # 지선
    lo_dongs = {}

    # 제8회 (2022)
    local8_dir = os.path.join(EXTRACT_DIR, "제8회_전국동시지방선거_읍면동별_개표결과-게시판게시")
    local8_files = [
        ("구시군장선거.xlsx", "시장", "2022 지선 (제8회) 시장"),
        ("시도지사선거.xlsx", "시도지사", "2022 지선 (제8회) 시도지사"),
        ("구시군의원선거.xlsx", "기초의원", "2022 지선 (제8회) 기초의원"),
        ("시도의원선거.xlsx", "도의원", "2022 지선 (제8회) 도의원"),
        ("교육감선거.xlsx", "교육감", "2022 지선 (제8회) 교육감"),
        ("기초비례의원선거.xlsx", "기초비례", "2022 지선 (제8회) 기초비례"),
        ("광역비례의원선거.xlsx", "광역비례", "2022 지선 (제8회) 광역비례"),
    ]

    for fname, sub_type, label in local8_files:
        filepath = os.path.join(local8_dir, fname)
        if os.path.exists(filepath):
            results = parse_local_election(filepath, sub_type, label, "2022.06.01")
            if results:
                lo_dongs[label] = results

    # 제7회 (2018)
    local7_dir = os.path.join(EXTRACT_DIR, "2018전국동시지방선거 개표결과(제7회)")
    local7_files = [
        ("20180619-7지선-02-(구시군의장)_읍면동별개표자료.xlsx", "시장", "2018 지선 (제7회) 시장"),
        ("20180619-7지선-01-(시도지사)_읍면동별개표자료.xlsx", "시도지사", "2018 지선 (제7회) 시도지사"),
        ("20180619-7지선-04-(구시군의회의원)_읍면동별개표자료.xlsx", "기초의원", "2018 지선 (제7회) 기초의원"),
        ("20180619-7지선-03-(시도의회의원)_읍면동별개표자료.xlsx", "도의원", "2018 지선 (제7회) 도의원"),
        ("20180619-7지선-07-(교육감)_읍면동별개표자료.xlsx", "교육감", "2018 지선 (제7회) 교육감"),
        ("20180619-7지선-06-(기초비례)_읍면동별개표자료.xlsx", "기초비례", "2018 지선 (제7회) 기초비례"),
        ("20180619-7지선-05-(광역비례)_읍면동별개표자료.xlsx", "광역비례", "2018 지선 (제7회) 광역비례"),
    ]

    for fname, sub_type, label in local7_files:
        filepath = os.path.join(local7_dir, fname)
        if os.path.exists(filepath):
            results = parse_local_2018(filepath, sub_type, label, "2018.06.13")
            if results:
                lo_dongs[label] = results

    # 병합
    print(f"\n{'='*50}")
    print("dongResults 병합 중...")

    el_file = os.path.join(OUTPUT_DIR, f"{CITY_CODE}-elections.json")
    lo_file = os.path.join(OUTPUT_DIR, f"{CITY_CODE}-local-elections.json")

    if os.path.exists(el_file) and el_dongs:
        merge_dong_results(el_file, el_dongs)
        for label, dongs in el_dongs.items():
            print(f"  일반: {label} → {len(dongs)}개 읍면동")

    if os.path.exists(lo_file) and lo_dongs:
        merge_dong_results(lo_file, lo_dongs)
        for label, dongs in lo_dongs.items():
            print(f"  지선: {label} → {len(dongs)}개 읍면동")

    print("\n완료!")


if __name__ == "__main__":
    main()
